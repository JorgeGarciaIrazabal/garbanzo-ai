"""Service for uploading, chunking, embedding, and searching knowledge-base documents."""

from __future__ import annotations

import asyncio
import io
import logging
import uuid
from dataclasses import dataclass

from sqlalchemy import delete, desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.db.session import async_session_maker
from app.models.knowledge_base import KnowledgeChunk, KnowledgeDocument
from app.services.embedding_provider import EmbeddingProvider, get_embedding_provider

logger = logging.getLogger(__name__)


@dataclass
class RetrievedChunk:
    document_id: str
    document_filename: str
    content: str
    score: float


# ---------------------------------------------------------------------------
# Text extraction helpers
# ---------------------------------------------------------------------------


def _extract_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    pages: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            pages.append(text)
    return "\n\n".join(pages)


def _extract_csv(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def _extract_spreadsheet(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        parts.append(f"## Sheet: {sheet_name}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            parts.append(" | ".join(cells))
    return "\n".join(parts)


def _extract_plain(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def _looks_like_text(text: str, sample_size: int = 4000) -> bool:
    """Heuristic check that extracted content is human-readable text.

    Binary input decoded with errors="replace" is dominated by U+FFFD
    replacement characters and control bytes; readable text is not.
    """
    sample = text[:sample_size]
    if not sample:
        return False
    bad = sum(
        1
        for ch in sample
        if ch == "�" or (ord(ch) < 32 and ch not in "\n\r\t")
    )
    return bad / len(sample) < 0.10


def extract_text(data: bytes, filename: str, mime_type: str) -> str:
    """Dispatch to the right extractor for ``mime_type`` / filename."""
    mime = (mime_type or "").lower()
    lower_name = filename.lower()

    if mime == "application/pdf" or lower_name.endswith(".pdf"):
        return _extract_pdf(data)
    if mime == "text/csv" or lower_name.endswith(".csv"):
        return _extract_csv(data)
    if (
        mime
        in {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "application/vnd.oasis.opendocument.spreadsheet",
        }
        or lower_name.endswith((".xlsx", ".xls", ".ods"))
    ):
        return _extract_spreadsheet(data)
    return _extract_plain(data)


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------


def chunk_text(text: str, chunk_size: int, overlap: int) -> list[str]:
    """Split ``text`` into overlapping character-based chunks.

    Prefers to break on paragraph / sentence boundaries when possible, with
    a hard fallback to a fixed-size slide when the chunk is too long.
    """
    text = text.strip()
    if not text:
        return []
    if overlap < 0:
        overlap = 0
    if overlap >= chunk_size:
        overlap = chunk_size // 4

    chunks: list[str] = []
    start = 0
    n = len(text)
    while start < n:
        end = min(start + chunk_size, n)
        # Try to break on paragraph or sentence end when not at the tail.
        if end < n:
            window = text[start:end]
            for sep in ("\n\n", "\n", ". ", " "):
                idx = window.rfind(sep)
                if idx >= chunk_size // 2:
                    end = start + idx + len(sep)
                    break
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= n:
            break
        start = max(end - overlap, start + 1)
    return chunks


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------


class KnowledgeBaseService:
    """CRUD + search + async embedding for a user's knowledge base."""

    def __init__(
        self,
        db: AsyncSession,
        embedding_provider: EmbeddingProvider | None = None,
        settings=None,
    ):
        self.db = db
        self._embedding_provider = embedding_provider or get_embedding_provider()
        self._settings = settings or get_settings()

    # ----- CRUD -----

    async def create_document(
        self,
        user_id: str,
        filename: str,
        mime_type: str,
        file_bytes: bytes,
    ) -> KnowledgeDocument:
        """Create a document row and synchronously extract + persist chunks.

        Embedding generation is dispatched to a background task so the caller
        does not block on the LLM.
        """
        text = extract_text(file_bytes, filename, mime_type).strip()
        if not text:
            raise ValueError("Could not extract any text from this file.")
        if not _looks_like_text(text):
            # MIME types are client-controlled, so binary files (executables,
            # images renamed .txt, …) can reach the plain-text fallback. The
            # mojibake they extract would pollute search results — validate
            # the content itself instead of trusting the declared type.
            raise ValueError(
                "File does not appear to contain readable text. "
                "Supported: PDF, CSV, spreadsheets, and plain-text files."
            )

        pieces = chunk_text(
            text,
            chunk_size=self._settings.kb_chunk_size,
            overlap=self._settings.kb_chunk_overlap,
        )
        if not pieces:
            raise ValueError("File produced no usable chunks.")

        doc = KnowledgeDocument(
            id=str(uuid.uuid4()),
            user_id=user_id,
            filename=filename,
            mime_type=mime_type or "",
            file_size=len(file_bytes),
            status="processing",
            chunk_count=len(pieces),
        )
        self.db.add(doc)
        for idx, piece in enumerate(pieces):
            self.db.add(
                KnowledgeChunk(
                    id=str(uuid.uuid4()),
                    document_id=doc.id,
                    user_id=user_id,
                    chunk_index=idx,
                    content=piece,
                )
            )
        await self.db.commit()
        await self.db.refresh(doc)

        if self._settings.kb_background_embedding:
            asyncio.create_task(self._embed_document_task(doc.id))
            logger.info(
                "KB: queued embedding for doc %s (%d chunks) user=%s",
                doc.id,
                len(pieces),
                user_id,
            )
        return doc

    async def list_documents(self, user_id: str) -> list[KnowledgeDocument]:
        query = (
            select(KnowledgeDocument)
            .where(KnowledgeDocument.user_id == user_id)
            .order_by(desc(KnowledgeDocument.created_at))
        )
        result = await self.db.execute(query)
        return list(result.scalars().all())

    async def get_document(
        self, document_id: str, user_id: str
    ) -> KnowledgeDocument | None:
        query = select(KnowledgeDocument).where(
            KnowledgeDocument.id == document_id,
            KnowledgeDocument.user_id == user_id,
        )
        return (await self.db.execute(query)).scalar_one_or_none()

    async def delete_document(self, document_id: str, user_id: str) -> bool:
        doc = await self.get_document(document_id, user_id)
        if not doc:
            return False
        # Delete chunks explicitly so we don't depend on DB-level cascade
        # (SQLite in tests doesn't enforce FKs by default).
        await self.db.execute(
            delete(KnowledgeChunk).where(KnowledgeChunk.document_id == document_id)
        )
        await self.db.delete(doc)
        await self.db.commit()
        logger.info("KB: deleted doc %s for user %s", document_id, user_id)
        return True

    # ----- Search -----

    async def search(
        self,
        user_id: str,
        query: str,
        limit: int | None = None,
    ) -> list[RetrievedChunk]:
        """Return the top-K chunks for ``query`` across the user's knowledge base.

        Hybrid retrieval: pgvector cosine similarity fused with Postgres
        full-text rank (``score = w*semantic + (1-w)*lexical``), so exact
        keyword queries can't be missed by embeddings alone. Falls back to
        semantic-only when full-text search is unavailable, and chunks below
        ``kb_min_score`` are dropped rather than injected as noise. Chunks
        without an embedding yet are skipped.
        """
        query = (query or "").strip()
        if not query:
            return []
        limit = limit or self._settings.kb_top_k

        try:
            embeddings = await self._embedding_provider.embed([query])
        except Exception:
            logger.exception("KB: failed to embed query")
            return []
        if not embeddings:
            return []
        query_vector = embeddings[0]

        chunks = await self._hybrid_search(user_id, query, query_vector, limit)
        if chunks is None:
            chunks = await self._semantic_search(user_id, query_vector, limit)

        kept = [c for c in chunks if c.score >= self._settings.kb_min_score]
        if len(kept) < len(chunks):
            logger.info(
                "KB: dropped %d of %d chunks below min score %.2f",
                len(chunks) - len(kept),
                len(chunks),
                self._settings.kb_min_score,
            )
        return kept

    def _base_chunk_query(self, user_id: str):
        return (
            select(
                KnowledgeChunk.document_id,
                KnowledgeChunk.content,
                KnowledgeDocument.filename,
            )
            .join(KnowledgeDocument, KnowledgeDocument.id == KnowledgeChunk.document_id)
            .where(
                KnowledgeChunk.user_id == user_id,
                KnowledgeChunk.embedding.isnot(None),
                KnowledgeDocument.status == "ready",
            )
        )

    async def _hybrid_search(
        self, user_id: str, query: str, query_vector: list[float], limit: int
    ) -> list[RetrievedChunk] | None:
        """Fused semantic + lexical retrieval. Returns None when the database
        can't run it (no pgvector / no full-text support), so the caller can
        fall back."""
        from sqlalchemy import Float, cast, func

        weight = self._settings.kb_semantic_weight
        try:
            semantic = 1.0 - KnowledgeChunk.embedding.cosine_distance(query_vector)
            # Normalization flag 32 maps ts_rank_cd into [0, 1) so it's
            # commensurable with cosine similarity.
            lexical = func.coalesce(
                func.ts_rank_cd(
                    func.to_tsvector("english", KnowledgeChunk.content),
                    func.websearch_to_tsquery("english", query),
                    32,
                ),
                0.0,
            )
            fused = (
                cast(semantic, Float) * weight
                + cast(lexical, Float) * (1.0 - weight)
            ).label("score")

            stmt = (
                self._base_chunk_query(user_id)
                .add_columns(fused)
                .order_by(fused.desc())
                .limit(limit)
            )
            # SAVEPOINT: a failed statement must not abort the caller's
            # transaction — chat streaming shares this session and has
            # un-committed work in flight (a session-level rollback here
            # would silently discard the user's message).
            async with self.db.begin_nested():
                rows = (await self.db.execute(stmt)).all()
        except Exception as e:
            logger.warning("KB: hybrid search unavailable (%s); semantic only", e)
            return None
        return [
            RetrievedChunk(
                document_id=row.document_id,
                document_filename=row.filename,
                content=row.content,
                score=float(row.score),
            )
            for row in rows
        ]

    async def _semantic_search(
        self, user_id: str, query_vector: list[float], limit: int
    ) -> list[RetrievedChunk]:
        """Cosine-only retrieval — the pre-hybrid behavior."""
        distance = KnowledgeChunk.embedding.cosine_distance(query_vector)
        stmt = (
            self._base_chunk_query(user_id)
            .add_columns(distance.label("distance"))
            .order_by(distance.asc())
            .limit(limit)
        )
        rows = (await self.db.execute(stmt)).all()
        return [
            RetrievedChunk(
                document_id=row.document_id,
                document_filename=row.filename,
                content=row.content,
                score=1.0 - float(row.distance),
            )
            for row in rows
        ]

    # ----- Embedding task -----

    async def _embed_document_task(self, document_id: str) -> None:
        """Background task: embed every chunk of a document.

        Runs in its own DB session so it survives the request that kicked it
        off.
        """
        try:
            async with async_session_maker() as session:
                doc = await session.get(KnowledgeDocument, document_id)
                if not doc:
                    return
                chunk_rows = (
                    await session.execute(
                        select(KnowledgeChunk)
                        .where(KnowledgeChunk.document_id == document_id)
                        .order_by(KnowledgeChunk.chunk_index)
                    )
                ).scalars().all()
                if not chunk_rows:
                    doc.status = "ready"
                    await session.commit()
                    return

                # Embed in batches to keep request size reasonable.
                batch_size = 16
                for i in range(0, len(chunk_rows), batch_size):
                    batch = chunk_rows[i : i + batch_size]
                    texts = [c.content for c in batch]
                    try:
                        vectors = await self._embedding_provider.embed(texts)
                    except Exception as exc:
                        logger.exception("KB: embedding batch failed for doc %s", document_id)
                        doc.status = "failed"
                        doc.error_message = str(exc)[:1000]
                        await session.commit()
                        return
                    for chunk, vec in zip(batch, vectors, strict=True):
                        chunk.embedding = vec
                    await session.flush()

                doc.status = "ready"
                doc.error_message = None
                await session.commit()
                logger.info("KB: embedded doc %s (%d chunks)", document_id, len(chunk_rows))
        except Exception:
            logger.exception("KB: background embedding task crashed for doc %s", document_id)
            try:
                async with async_session_maker() as session:
                    doc = await session.get(KnowledgeDocument, document_id)
                    if doc:
                        doc.status = "failed"
                        doc.error_message = "Unexpected error during embedding."
                        await session.commit()
            except Exception:
                logger.exception("KB: failed to mark doc %s as failed", document_id)

    # ----- Maintenance -----

    async def purge_orphan_chunks(self, user_id: str) -> int:
        """Remove chunks whose document row is missing (defensive cleanup)."""
        result = await self.db.execute(
            delete(KnowledgeChunk).where(
                KnowledgeChunk.user_id == user_id,
                ~select(KnowledgeDocument.id)
                .where(KnowledgeDocument.id == KnowledgeChunk.document_id)
                .exists(),
            )
        )
        await self.db.commit()
        return result.rowcount or 0

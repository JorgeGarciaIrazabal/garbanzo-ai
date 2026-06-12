"""Service for sending messages and streaming LLM responses."""

import asyncio
import base64
import io
import json
import logging
import re
import uuid
from collections.abc import AsyncIterator
from typing import Any

from sqlalchemy import delete, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.models.message import Message
from app.schemas.chat import AttachmentIn, ChatOptions, ModelInfo
from app.services.conversation_service import ConversationService
from app.services.knowledge_base_service import KnowledgeBaseService
from app.services.llm_provider import (
    ChatChunk,
    LLMProvider,
    ProviderRegistry,
    resolve_context_length,
)
from app.services.llm_provider import Message as LLMMessage
from app.services.mcp_service import (
    MCPService,
    build_tool_payload,
    split_tool_key,
    tool_key,
)
from app.services.memory_service import MemoryService
from app.services.ollama_provider import OllamaProvider
from app.services.system_prompt_service import SystemPromptService
from app.services.token_counter import get_token_counter

MAX_TOOL_ITERATIONS = 5

logger = logging.getLogger(__name__)


def _stringify_tool_result(result: Any) -> str:
    """Coerce a tool-call result dict into a string suitable for ``Message.content``."""
    try:
        return json.dumps(result, default=str)
    except Exception:
        return str(result)


def _maybe_parse_inline_tool_calls(content: str) -> list[dict[str, Any]] | None:
    """If ``content`` is a JSON list of tool-call objects, return them.

    Defensive cleanup for legacy assistant messages that contain raw
    tool-call JSON in their content (an artifact of an earlier bug where
    the call list was stringified into ``content``). Returning the parsed
    calls lets the history-replay path send them via the proper
    ``tool_calls`` field instead of as imitable text.
    """
    stripped = content.strip()
    if not stripped or stripped[0] != "[":
        return None
    try:
        parsed = json.loads(stripped)
    except (ValueError, TypeError):
        return None
    if not isinstance(parsed, list) or not parsed:
        return None
    for item in parsed:
        if not isinstance(item, dict):
            return None
        if "name" not in item or "arguments" not in item:
            return None
    return parsed


_TITLE_PROMPT = (
    "Write a short title (3-5 words) for this conversation. "
    "Reply with ONLY the title — no quotes, no trailing punctuation, "
    "no explanation.\n\nUser: {user}\nAssistant: {assistant}"
)


def _clean_generated_title(raw: str) -> str:
    """Normalize an LLM-generated title to a single clean line."""
    text = raw.strip()
    # Defensive: some models inline reasoning despite the prompt.
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()
    first_line = next((ln.strip() for ln in text.splitlines() if ln.strip()), "")
    return first_line.strip("\"'`#* ").rstrip(".!,:;")[:60]


async def generate_conversation_title(
    provider: LLMProvider, model: str, user_text: str, assistant_text: str
) -> str:
    """Generate a 3-5 word conversation title; returns "" on empty output."""
    prompt = _TITLE_PROMPT.format(
        user=user_text[:500], assistant=assistant_text[:500]
    )
    raw = await provider.chat(
        messages=[LLMMessage(role="user", content=prompt)],
        model=model,
        options=ChatOptions(temperature=0.3, max_tokens=30),
    )
    return _clean_generated_title(raw)


async def _generate_title_task(
    provider: LLMProvider, conversation_id: str, model: str,
    user_text: str, assistant_text: str,
) -> None:
    """Background task: generate and persist a title with its own session."""
    try:
        title = await generate_conversation_title(
            provider, model, user_text, assistant_text
        )
        if not title:
            return
        from app.db.session import async_session_maker
        from app.models.conversation import Conversation

        async with async_session_maker() as db:
            conversation = await db.get(Conversation, conversation_id)
            if conversation is not None:
                conversation.title = title
                await db.commit()
        logger.info("Auto-titled conversation %s: %r", conversation_id, title)
    except Exception as e:
        logger.warning("Title generation failed for %s: %s", conversation_id, e)


async def _extract_pdf_text(data: str) -> str:
    """Extract text from a base64-encoded PDF."""
    try:
        from pypdf import PdfReader

        pdf_bytes = base64.b64decode(data)
        reader = PdfReader(io.BytesIO(pdf_bytes))
        text_pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_pages.append(text)
        return "\n".join(text_pages) if text_pages else "[PDF: no extractable text]"
    except Exception as e:
        logger.warning("PDF extraction failed: %s", e)
        return f"[PDF extraction error: {e}]"


async def _extract_csv_text(data: str, filename: str) -> str:
    """Extract and summarize CSV content."""
    text_bytes = base64.b64decode(data)
    text_content = text_bytes.decode("utf-8", errors="replace")
    lines = text_content.splitlines()
    row_count = len(lines)
    summary = f"[CSV: {filename}, {row_count} rows]\n"
    if row_count <= 50:
        summary += text_content
    else:
        summary += "--- Preview (first 50 rows) ---\n"
        summary += "\n".join(lines[:50])
        summary += f"\n... and {row_count - 50} more rows"
    return summary


async def _extract_spreadsheet_text(data: str, filename: str) -> str:
    """Extract and summarize Excel/openoffice spreadsheet content."""
    try:
        from openpyxl import load_workbook

        xls_bytes = base64.b64decode(data)
        wb = load_workbook(filename=io.BytesIO(xls_bytes), read_only=True, data_only=True)
        summary_parts = []
        total_rows = 0
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            total_rows += len(rows)
            if rows:
                summary_parts.append(f"Sheet: {sheet_name}")
                if len(rows) <= 20:
                    for row in rows:
                        cells = [str(c) if c is not None else "" for c in row]
                        summary_parts.append(" | ".join(cells))
                else:
                    summary_parts.append(f"  {len(rows)} rows")
                    for row in rows[:10]:
                        cells = [str(c) if c is not None else "" for c in row]
                        summary_parts.append(" | ".join(cells))
                    summary_parts.append(f"  ... and {len(rows) - 10} more rows")
            summary_parts.append("")
        return f"[Spreadsheet: {filename}, {total_rows} total rows]\n" + "\n".join(summary_parts)
    except Exception as e:
        logger.warning("Spreadsheet extraction failed: %s", e)
        return f"[Spreadsheet extraction error: {e}]"


class ChatService:
    """Handles sending messages and streaming LLM responses.

    Conversation CRUD is delegated to ``ConversationService``.
    """

    _active_streams: dict[str, asyncio.Event] = {}

    def __init__(self, db: AsyncSession, *, provider_name: str = "ollama"):
        self.db = db
        self._provider_name = provider_name
        self._conversations = ConversationService(db)
        self._memories = MemoryService(db)
        self._system_prompts = SystemPromptService(db)
        self._mcp = MCPService(db)
        self._kb = KnowledgeBaseService(db)
        self._ensure_default_provider()

    @classmethod
    def cancel_stream(cls, conversation_id: str) -> bool:
        """Signal an active stream to stop. Returns True if found."""
        event = cls._active_streams.get(conversation_id)
        if event:
            event.set()
            return True
        return False

    @property
    def conversations(self) -> ConversationService:
        return self._conversations

    def _ensure_default_provider(self) -> None:
        if "ollama" not in ProviderRegistry.list_providers():
            settings = get_settings()
            ProviderRegistry.register(OllamaProvider(base_url=settings.ollama_base_url))

    def _get_provider(self) -> LLMProvider:
        provider = ProviderRegistry.get(self._provider_name)
        if provider is None:
            raise ValueError(f"Unknown provider: {self._provider_name}")
        return provider

    async def _get_context_length(self, model: str) -> int:
        """Effective context window for ``model`` — see resolve_context_length."""
        return await resolve_context_length(self._get_provider(), model)

    async def _maybe_summarize_context(
        self,
        conversation,
        messages: list,
    ) -> None:
        """Summarize older messages if context window is near full (>80%)."""
        if not messages:
            return

        # Prefer the provider-reported prompt size from the last turn (exact);
        # fall back to a TokenCounter estimate so conversations without one
        # (first turns, imported history) still get summarized in time.
        last_assistant = next(
            (m for m in reversed(messages) if m.role == "assistant" and m.meta),
            None,
        )
        tokens_prompt = None
        if last_assistant and last_assistant.meta:
            tokens_prompt = last_assistant.meta.get("tokens_prompt")
        if not tokens_prompt:
            tokens_prompt = get_token_counter().count_messages(
                [m.content or "" for m in messages]
            )

        context_length = await self._get_context_length(conversation.model)
        if tokens_prompt < int(context_length * 0.8):
            return

        # Find start index (respect existing summary boundary)
        start_idx = 0
        if conversation.context_summary_until_id:
            for i, m in enumerate(messages):
                if m.id == conversation.context_summary_until_id:
                    start_idx = i + 1
                    break

        # Keep last 10 messages intact, summarize everything before
        end_idx = max(start_idx, len(messages) - 10)
        if end_idx <= start_idx:
            return

        messages_to_summarize = messages[start_idx:end_idx]

        text_parts = []
        for msg in messages_to_summarize:
            if msg.role == "tool_call":
                # The content is raw JSON; a name list reads better and
                # won't teach the summarizer to emit tool-call syntax.
                calls = (msg.meta or {}).get("tool_calls") or []
                names = ", ".join(c.get("name", "?") for c in calls) or "unknown"
                text_parts.append(f"[Called tools: {names}]")
            elif msg.role == "tool_result":
                name = (msg.meta or {}).get("tool_name", "tool")
                text_parts.append(f"[Result from {name}]: {msg.content[:300]}")
            else:
                text_parts.append(f"{msg.role.upper()}: {msg.content[:800]}")
        summary_input = "\n\n".join(text_parts)

        prompt = (
            "Condense the following conversation excerpt into rolling context "
            "notes (at most 8 sentences, plain prose). You MUST preserve:\n"
            "- the user's goals, preferences, and constraints\n"
            "- key facts, names, numbers, and decisions made\n"
            "- important results returned by tools\n"
            "- any open questions or unfinished work\n"
            "Omit pleasantries and repetition. Do not address the user; "
            "write neutral notes.\n\n"
            + summary_input
        )

        try:
            provider = self._get_provider()
            summary_parts: list[str] = []
            # The summarize input is by definition ~80% of the window —
            # without num_ctx the request would run at the runtime default
            # (typically 4096) and silently truncate exactly what we're
            # trying to preserve.
            async for chunk in provider.stream_chat(
                messages=[LLMMessage(role="user", content=prompt)],
                model=conversation.model,
                options=ChatOptions(temperature=0.3, num_ctx=context_length),
            ):
                if not chunk.is_thinking and chunk.content:
                    summary_parts.append(chunk.content)

            new_summary = "".join(summary_parts).strip()
            if not new_summary:
                return

            if conversation.context_summary:
                new_summary = f"{conversation.context_summary}\n\n{new_summary}"

            conversation.context_summary = new_summary
            conversation.context_summary_until_id = messages_to_summarize[-1].id
            await self.db.flush()

        except Exception as e:
            logger.warning("Auto-summarization failed: %s", e)

    async def send_message(
        self,
        conversation_id: str,
        user_id: str,
        content: str,
        options: ChatOptions | None = None,
        attachments: list[AttachmentIn] | None = None,
    ) -> AsyncIterator[ChatChunk]:
        """Save user message, stream LLM response, and persist the result."""
        conversation = await self._conversations.get(conversation_id, user_id)
        if not conversation:
            yield ChatChunk(
                content="Conversation not found",
                is_finished=True,
                metadata={"error": True, "error_type": "not_found"},
            )
            return

        # Build the content that goes into the DB (append document text inline)
        stored_content = content
        attachment_meta: list[dict] = []
        image_b64_list: list[str] = []

        if attachments:
            doc_texts: list[str] = []
            for att in attachments:
                attachment_meta.append(
                    {"name": att.name, "mime_type": att.mime_type, "type": att.type}
                )
                if att.type == "image":
                    image_b64_list.append(att.data)
                elif att.type == "document" and att.data:
                    # Try to extract text from PDFs and spreadsheets
                    mime = att.mime_type.lower()
                    filename = att.name
                    extracted_text = att.data  # fallback: use raw text as-is

                    if mime == "application/pdf":
                        extracted_text = await _extract_pdf_text(att.data)
                    elif mime == "text/csv":
                        extracted_text = await _extract_csv_text(att.data, filename)
                    elif mime in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  "application/vnd.ms-excel",
                                  "application/vnd.oasis.opendocument.spreadsheet"):
                        extracted_text = await _extract_spreadsheet_text(att.data, filename)
                    elif mime.startswith("text/") or filename.endswith(('.txt', '.md', '.json', '.py', '.js', '.ts', '.dart', '.csv')):
                        # Plain text files - decode from base64
                        try:
                            extracted_text = base64.b64decode(att.data).decode("utf-8", errors="replace")
                        except Exception:
                            extracted_text = att.data

                    doc_texts.append(f"[Attached file: {filename}]\n{extracted_text}")

            if doc_texts:
                stored_content = content + "\n\n" + "\n\n".join(doc_texts)

        user_message = Message(
            id=str(uuid.uuid4()),
            conversation_id=conversation_id,
            role="user",
            content=stored_content,
            meta={"attachments": attachment_meta} if attachment_meta else None,
            conversation=conversation,
        )
        self.db.add(user_message)
        await self.db.flush()

        conversation.updated_at = func.now()  # type: ignore[assignment]
        await self.db.flush()

        async for chunk in self._stream_assistant_turn(
            conversation=conversation,
            options=options,
            pending_images=image_b64_list,
        ):
            yield chunk

    async def regenerate_message(
        self,
        conversation_id: str,
        user_id: str,
        message_id: str,
        options: ChatOptions | None = None,
    ) -> AsyncIterator[ChatChunk]:
        """Re-run the LLM from the user message preceding ``message_id``.

        ``message_id`` should identify an assistant message. All messages from
        that message onward (assistant + any trailing tool_call/tool_result)
        are deleted, and the response is re-streamed based on the remaining
        history.
        """
        conversation = await self._conversations.get(
            conversation_id, user_id, include_messages=True
        )
        if not conversation:
            yield ChatChunk(
                content="Conversation not found",
                is_finished=True,
                metadata={"error": True, "error_type": "not_found"},
            )
            return

        messages = list(conversation.messages) if conversation.messages else []
        target = next((m for m in messages if m.id == message_id), None)
        if not target:
            yield ChatChunk(
                content="Message not found",
                is_finished=True,
                metadata={"error": True, "error_type": "not_found"},
            )
            return
        if target.role != "assistant":
            yield ChatChunk(
                content="Can only regenerate assistant messages",
                is_finished=True,
                metadata={"error": True, "error_type": "invalid_role"},
            )
            return

        # Delete the target message and everything after it.
        target_index = next(i for i, m in enumerate(messages) if m.id == message_id)
        ids_to_delete = [m.id for m in messages[target_index:]]
        await self._delete_messages_by_ids(ids_to_delete)
        await self.db.flush()
        await self.db.refresh(conversation, attribute_names=["messages"])

        conversation.updated_at = func.now()  # type: ignore[assignment]
        await self.db.flush()

        async for chunk in self._stream_assistant_turn(
            conversation=conversation,
            options=options,
        ):
            yield chunk

    async def edit_and_resend(
        self,
        conversation_id: str,
        user_id: str,
        message_id: str,
        new_content: str,
        options: ChatOptions | None = None,
    ) -> AsyncIterator[ChatChunk]:
        """Edit a user message and re-run the conversation from that point.

        Updates the message's text (attachments are preserved) and deletes
        every message that came after it, then streams a fresh assistant
        response.
        """
        conversation = await self._conversations.get(
            conversation_id, user_id, include_messages=True
        )
        if not conversation:
            yield ChatChunk(
                content="Conversation not found",
                is_finished=True,
                metadata={"error": True, "error_type": "not_found"},
            )
            return

        messages = list(conversation.messages) if conversation.messages else []
        target = next((m for m in messages if m.id == message_id), None)
        if not target:
            yield ChatChunk(
                content="Message not found",
                is_finished=True,
                metadata={"error": True, "error_type": "not_found"},
            )
            return
        if target.role != "user":
            yield ChatChunk(
                content="Can only edit user messages",
                is_finished=True,
                metadata={"error": True, "error_type": "invalid_role"},
            )
            return

        # Preserve the original attachment text block (everything after the
        # first [Attached file: ...] marker) so re-sending keeps the docs.
        appended_block = ""
        marker = "\n\n[Attached file: "
        idx = target.content.find(marker)
        if idx >= 0:
            appended_block = target.content[idx:]
        target.content = new_content + appended_block

        # Remove everything after this message.
        target_index = next(i for i, m in enumerate(messages) if m.id == message_id)
        ids_to_delete = [m.id for m in messages[target_index + 1 :]]
        await self._delete_messages_by_ids(ids_to_delete)
        await self.db.flush()
        await self.db.refresh(conversation, attribute_names=["messages"])

        conversation.updated_at = func.now()  # type: ignore[assignment]
        await self.db.flush()

        # Rehydrate images from the message metadata so the LLM still sees them.
        image_b64_list: list[str] = []
        raw_attachments = (target.meta or {}).get("attachments") or []
        for att in raw_attachments:
            if isinstance(att, dict) and att.get("type") == "image" and att.get("data"):
                image_b64_list.append(att["data"])

        async for chunk in self._stream_assistant_turn(
            conversation=conversation,
            options=options,
            pending_images=image_b64_list or None,
        ):
            yield chunk

    async def _delete_messages_by_ids(self, message_ids: list[str]) -> None:
        """Delete messages matching any of the provided IDs."""
        if not message_ids:
            return
        await self.db.execute(delete(Message).where(Message.id.in_(message_ids)))

    async def _stream_assistant_turn(
        self,
        conversation,
        options: ChatOptions | None = None,
        pending_images: list[str] | None = None,
    ) -> AsyncIterator[ChatChunk]:
        """Stream an LLM response for the current state of ``conversation``.

        Assumes the conversation's messages list already reflects the desired
        history (e.g. user turn appended, or trailing assistant trimmed).
        Handles summarization, tool-call iterations, and persistence.
        """
        conversation_id = conversation.id
        # Captured before streaming: a turn with no prior assistant message
        # is the conversation's first exchange and gets an auto-title.
        existing_messages = list(conversation.messages)
        is_first_exchange = not any(m.role == "assistant" for m in existing_messages)
        last_user_text = next(
            (m.content for m in reversed(existing_messages) if m.role == "user"),
            "",
        )

        await self._maybe_summarize_context(conversation, existing_messages)

        llm_messages, context_stats = await self._build_message_history_with_system_prompt(
            conversation.messages,
            pending_images,
            conversation.user_id,
            use_memory=conversation.use_memory,
            use_knowledge_base=conversation.use_knowledge_base,
            context_summary=conversation.context_summary,
            context_summary_until_id=conversation.context_summary_until_id,
            conversation_system_prompt=conversation.system_prompt,
        )

        provider = self._get_provider()
        opts = options or ChatOptions()

        # Allocate the effective window explicitly — without num_ctx, Ollama
        # runs at its own default (typically 4096) no matter what the model
        # supports, silently truncating long conversations. The server-side
        # value is also a ceiling: a client-supplied num_ctx may shrink the
        # window but never grow it (an oversized request would make the
        # runtime allocate an arbitrarily large KV cache).
        context_length = await self._get_context_length(conversation.model)
        opts.num_ctx = min(opts.num_ctx or context_length, context_length)

        cancel_event = asyncio.Event()
        ChatService._active_streams[conversation_id] = cancel_event

        try:
            ollama_tools, tool_lookup = await self._resolve_tools_for_conversation(
                conversation
            )

            # Thinking from tool-calling iterations (which produce no
            # `content` and so don't get persisted on their own) is carried
            # forward and attached to the next assistant message that does
            # have content. This way the persisted state is one tidy
            # assistant message per turn with cumulative reasoning, instead
            # of orphaned thinking + a separate answer message.
            pending_thinking: list[str] = []

            for _ in range(MAX_TOOL_ITERATIONS):
                full_response = ""
                thinking_content = ""
                metadata: dict | None = None
                tool_calls_this_iter: list[dict] | None = None

                async for chunk in provider.stream_chat(
                    messages=llm_messages,
                    model=conversation.model,
                    options=opts,
                    cancel_event=cancel_event,
                    tools=ollama_tools or None,
                ):
                    if chunk.tool_calls:
                        tool_calls_this_iter = chunk.tool_calls
                        yield chunk
                        continue
                    if chunk.is_thinking:
                        thinking_content += chunk.content
                    elif chunk.content:
                        full_response += chunk.content
                    if chunk.is_finished:
                        # Stamp the window we allocated so the persisted
                        # message meta and the live stream both carry the
                        # real denominator for context-usage display, plus
                        # what personal context informed this reply.
                        if chunk.metadata is None:
                            chunk.metadata = {}
                        chunk.metadata.setdefault("context_length", opts.num_ctx)
                        if context_stats.get("memories_used"):
                            chunk.metadata.setdefault(
                                "memories_used", context_stats["memories_used"]
                            )
                        if context_stats.get("kb_chunks_used"):
                            chunk.metadata.setdefault(
                                "kb_chunks_used", context_stats["kb_chunks_used"]
                            )
                        metadata = chunk.metadata
                    yield chunk

                if thinking_content:
                    pending_thinking.append(thinking_content)

                if full_response:
                    msg_meta = dict(metadata) if metadata else {}
                    combined_thinking = "\n\n".join(pending_thinking)
                    if combined_thinking:
                        msg_meta["thinking"] = combined_thinking
                    pending_thinking = []
                    assistant_message = Message(
                        id=str(uuid.uuid4()),
                        conversation_id=conversation_id,
                        role="assistant",
                        content=full_response,
                        meta=msg_meta or None,
                        # Keep the in-session relationship collection in sync
                        # (raw FK writes don't update conversation.messages),
                        # so a later turn on the same session sees this one.
                        conversation=conversation,
                    )
                    self.db.add(assistant_message)
                    await self.db.flush()
                    llm_messages.append(
                        LLMMessage(role="assistant", content=full_response)
                    )

                if not tool_calls_this_iter:
                    await self.db.commit()
                    if is_first_exchange and full_response and last_user_text:
                        self._spawn_title_generation(
                            conversation_id,
                            conversation.model,
                            last_user_text,
                            full_response,
                        )
                    return

                tool_call_payload = json.dumps(tool_calls_this_iter)
                tool_call_msg = Message(
                    id=str(uuid.uuid4()),
                    conversation_id=conversation_id,
                    role="tool_call",
                    content=tool_call_payload,
                    meta={"tool_calls": tool_calls_this_iter},
                )
                self.db.add(tool_call_msg)
                await self.db.flush()

                # Send tool calls via the API's native field — never as raw
                # JSON in content, or the model mimics the format and emits
                # the call as text on subsequent turns.
                llm_messages.append(
                    LLMMessage(
                        role="assistant",
                        content="",
                        tool_calls=tool_calls_this_iter,
                    )
                )

                for call in tool_calls_this_iter:
                    result = await self._execute_tool_call(call, tool_lookup)
                    result_meta = {
                        "tool_call_id": call.get("id"),
                        "tool_name": call.get("name"),
                        "result": result,
                    }
                    result_text = _stringify_tool_result(result)
                    tool_result_msg = Message(
                        id=str(uuid.uuid4()),
                        conversation_id=conversation_id,
                        role="tool_result",
                        content=result_text,
                        meta=result_meta,
                    )
                    self.db.add(tool_result_msg)
                    await self.db.flush()

                    yield ChatChunk(
                        content="",
                        is_finished=False,
                        metadata={"tool_result": result_meta},
                    )

                    llm_messages.append(LLMMessage(role="tool", content=result_text))

                await self.db.commit()

            yield ChatChunk(
                content="",
                is_finished=True,
                metadata={
                    "error": True,
                    "error_type": "tool_iteration_cap",
                    "max_iterations": MAX_TOOL_ITERATIONS,
                },
            )

        except Exception as e:
            logger.exception("Error in chat streaming")
            yield ChatChunk(
                content=f"Error: {e}",
                is_finished=True,
                metadata={"error": True, "error_type": "streaming_error"},
            )
            await self.db.rollback()
        finally:
            ChatService._active_streams.pop(conversation_id, None)

    def _spawn_title_generation(
        self,
        conversation_id: str,
        model: str,
        user_text: str,
        assistant_text: str,
    ) -> None:
        """Fire-and-forget title generation on the conversation's model.

        Uses the already-loaded conversation model rather than a separate
        small model so low-RAM deployments don't pay a second model load.
        """
        asyncio.create_task(
            _generate_title_task(
                self._get_provider(),
                conversation_id,
                model,
                user_text,
                assistant_text,
            )
        )

    async def _resolve_tools_for_conversation(
        self, conversation
    ) -> tuple[list[dict], dict[str, tuple[str, str]]]:
        """Return ``(ollama_tools, lookup)`` for this conversation.

        ``lookup`` maps the function name advertised to the LLM back to
        ``(server_id, tool_name)`` so the executor can resolve calls.
        """
        enabled = getattr(conversation, "enabled_tools", None)
        if enabled is not None and not enabled:
            # [] → opt out of all tools
            return [], {}

        try:
            all_tools = await self._mcp.list_all_tools(enabled_only=True)
        except Exception as exc:
            logger.warning("Failed to list MCP tools: %s", exc)
            return [], {}

        if enabled is None:
            filtered = all_tools
        else:
            allowed = set(enabled)
            filtered = [
                t for t in all_tools if tool_key(t["server_id"], t["name"]) in allowed
            ]
        return build_tool_payload(filtered)

    async def _execute_tool_call(
        self, call: dict, lookup: dict[str, tuple[str, str]]
    ) -> dict:
        """Run a single tool call through ``MCPService``.

        ``call["name"]`` is the function name we advertised to the LLM. We
        resolve it back to ``(server_id, tool_name)`` via ``lookup``. As a
        legacy fallback we also accept the raw ``"server_id:tool_name"``
        form some older flows produced.
        """
        name = call.get("name") or ""
        args = call.get("arguments") or {}
        if isinstance(args, str):
            try:
                args = json.loads(args)
            except Exception:
                args = {}

        target = lookup.get(name)
        if target is None and ":" in name:
            target = split_tool_key(name)
        if target is None:
            return {"ok": False, "error": f"Unknown tool: {name}"}

        server_id, tool_name = target
        try:
            return await self._mcp.call_tool(server_id, tool_name, args)
        except Exception as exc:
            logger.exception("Tool execution failed: %s", name)
            return {"ok": False, "error": str(exc)}

    def _build_message_history(
        self,
        messages: list[Message],
        pending_images: list[str] | None = None,
    ) -> list[LLMMessage]:
        result: list[LLMMessage] = []
        for i, msg in enumerate(messages):
            is_last = i == len(messages) - 1
            images = pending_images if (is_last and pending_images) else None
            result.append(LLMMessage(role=msg.role, content=msg.content, images=images))
        return result

    async def _build_message_history_with_system_prompt(
        self,
        messages: list[Message],
        pending_images: list[str] | None = None,
        user_id: str | None = None,
        use_memory: bool = True,
        use_knowledge_base: bool = True,
        context_summary: str | None = None,
        context_summary_until_id: str | None = None,
        conversation_system_prompt: str | None = None,
    ) -> tuple[list[LLMMessage], dict[str, int]]:
        """Build message history with an optional system prompt prepended.

        If use_memory is True and user_id is provided, fetches relevant memories
        and prepends them to the system prompt.

        Args:
            messages: List of conversation messages
            pending_images: Base64-encoded images for multimodal context
            user_id: The user ID to fetch memories for
            use_memory: Whether to inject memories into the system prompt
            context_summary: Summary of earlier messages that were trimmed
            context_summary_until_id: ID of last message included in the summary

        Returns:
            List of LLMMessage with system prompt as first message if memories exist
        """
        result: list[LLMMessage] = []

        # Filter to unsummarized messages if summary exists
        filtered_messages = list(messages)
        if context_summary and context_summary_until_id:
            for i, m in enumerate(filtered_messages):
                if m.id == context_summary_until_id:
                    filtered_messages = filtered_messages[i + 1 :]
                    break

        # Latest user message drives knowledge-base retrieval.
        last_user_query: str | None = None
        for m in reversed(filtered_messages):
            if m.role == "user" and m.content:
                last_user_query = m.content
                break

        # Build system prompt from (conversation || global default) + memories + KB
        system_prompt, context_stats = await self._build_system_prompt(
            user_id=user_id,
            use_memory=use_memory,
            use_knowledge_base=use_knowledge_base,
            rag_query=last_user_query,
            conversation_system_prompt=conversation_system_prompt,
        )

        if system_prompt:
            result.append(LLMMessage(role="system", content=system_prompt))

        if context_summary:
            result.append(
                LLMMessage(
                    role="system",
                    content=f"[Earlier conversation summary]\n{context_summary}",
                )
            )

        # Add conversation messages
        for i, msg in enumerate(filtered_messages):
            is_last = i == len(filtered_messages) - 1
            images = pending_images if (is_last and pending_images) else None
            # Normalize tool-related roles for the LLM: Ollama expects "tool"
            # for a tool result; tool_call messages from the assistant collapse
            # back to an assistant turn carrying structured tool_calls. We
            # MUST NOT inline the tool-call JSON into content — the model
            # would imitate that format and emit raw JSON on the next turn.
            if msg.role == "tool_result":
                result.append(LLMMessage(role="tool", content=msg.content))
            elif msg.role == "tool_call":
                tool_calls = (msg.meta or {}).get("tool_calls") or []
                result.append(
                    LLMMessage(
                        role="assistant",
                        content="",
                        tool_calls=tool_calls or None,
                    )
                )
            elif msg.role == "assistant":
                # Legacy hygiene: if a past assistant turn was persisted with
                # raw tool-call JSON in content, recover it as structured
                # tool_calls so the model doesn't see (and imitate) the JSON.
                inline = _maybe_parse_inline_tool_calls(msg.content)
                if inline is not None:
                    result.append(
                        LLMMessage(role="assistant", content="", tool_calls=inline)
                    )
                else:
                    result.append(
                        LLMMessage(role=msg.role, content=msg.content, images=images)
                    )
            else:
                result.append(LLMMessage(role=msg.role, content=msg.content, images=images))

        return result, context_stats

    async def _build_system_prompt(
        self,
        user_id: str | None = None,
        use_memory: bool = True,
        use_knowledge_base: bool = True,
        rag_query: str | None = None,
        conversation_system_prompt: str | None = None,
    ) -> tuple[str, dict[str, int]]:
        """Build the system prompt from (conversation || user default) + memories.

        Preference order for the base prompt:
          1. ``conversation_system_prompt`` if set
          2. The user's ``default_system_prompt`` if set
          3. A generic fallback only if memories will be injected.

        Memories (when enabled and present) are appended after the base prompt.

        Returns ``(prompt, stats)`` where stats counts the context actually
        injected ({"memories_used": n, "kb_chunks_used": n}) so the client
        can show the user what informed the reply.
        """
        stats = {"memories_used": 0, "kb_chunks_used": 0}
        base_prompt = (conversation_system_prompt or "").strip()

        if not base_prompt and user_id:
            try:
                user_default = await self._system_prompts.get_user_default_prompt(user_id)
                if user_default:
                    base_prompt = user_default.strip()
            except Exception as e:
                logger.warning("Failed to load user default system prompt: %s", e)

        settings = get_settings()
        counter = get_token_counter()

        memory_block = ""
        if user_id and use_memory:
            try:
                # Ranked by semantic relevance to the current message (with
                # recency fallback), capped at top-K — not the whole store.
                memories = await self._memories.get_relevant_memories(
                    user_id=user_id,
                    query=rag_query,
                    limit=settings.memory_top_k,
                )
                if memories:
                    # Keep memories in list order until the budget runs out, so
                    # a large memory store can't crowd out the conversation.
                    budget = settings.memory_token_budget
                    used = 0
                    kept = []
                    for memory in memories:
                        cost = counter.count_text(memory.content) + 2
                        if kept and used + cost > budget:
                            break
                        used += cost
                        kept.append(memory)
                    if len(kept) < len(memories):
                        logger.info(
                            "Memory block over %d-token budget: injecting %d of %d memories",
                            budget,
                            len(kept),
                            len(memories),
                        )
                    lines = ["Relevant memories about the user:"]
                    for memory in kept:
                        lines.append(f"- {memory.content}")
                    lines.append("")
                    lines.append("Use these memories to personalize your responses.")
                    memory_block = "\n".join(lines)
                    stats["memories_used"] = len(kept)
            except Exception as e:
                logger.warning("Failed to load memories for system prompt: %s", e)

        kb_block = ""
        if user_id and use_knowledge_base and rag_query:
            try:
                matches = await self._kb.search(user_id=user_id, query=rag_query)
                if matches:
                    budget = settings.kb_token_budget
                    used = 0
                    lines = [
                        "Relevant excerpts from the user's knowledge base. "
                        "Cite the source filename when you use them; if none "
                        "are relevant, ignore this section.",
                        "",
                    ]
                    injected = 0
                    for m in matches:
                        cost = counter.count_text(m.content) + 10
                        if injected and used + cost > budget:
                            logger.info(
                                "KB block over %d-token budget: injecting %d of %d chunks",
                                budget,
                                injected,
                                len(matches),
                            )
                            break
                        used += cost
                        injected += 1
                        lines.append(
                            f"--- Source: {m.document_filename} "
                            f"(relevance: {m.score:.2f}) ---"
                        )
                        lines.append(m.content)
                        lines.append("")
                    kb_block = "\n".join(lines).rstrip()
                    stats["kb_chunks_used"] = injected
            except Exception as e:
                logger.warning("Failed to load KB context for system prompt: %s", e)

        if not base_prompt and not memory_block and not kb_block:
            return "", stats

        if not base_prompt and (memory_block or kb_block):
            base_prompt = "You are a helpful AI assistant."

        sections = [base_prompt]
        if memory_block:
            sections.append(memory_block)
        if kb_block:
            sections.append(kb_block)
        return "\n\n".join(sections), stats

    async def list_available_models(self) -> list[ModelInfo]:
        provider = self._get_provider()
        models = await provider.list_models()

        return [
            ModelInfo(
                id=m.id,
                name=m.name,
                description=m.description,
                context_length=m.context_length,
                provider=self._provider_name,
                supports_tools=m.supports_tools,
                supports_vision=m.supports_vision,
                supports_thinking=m.supports_thinking,
            )
            for m in models
        ]

    async def health_check(self) -> dict[str, bool]:
        results = {}
        for name in ProviderRegistry.list_providers():
            provider = ProviderRegistry.get(name)
            if provider:
                results[name] = await provider.health_check()
        return results

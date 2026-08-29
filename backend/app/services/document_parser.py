"""Text extraction for document attachments sent alongside a chat message.

Handles PDFs, CSVs, spreadsheets (xlsx/xls/ods), and plain-text files.
Everything funnels through :func:`extract_attachment_text`, which dispatches
on ``attachment.mime_type`` (falling back to the filename extension for
plain text) and never raises — extraction failures degrade to an inline
error marker or the raw attachment data so a single bad upload can't take
down message sending.

Attachment bytes are base64 on the current wire contract. Payloads without an
explicit encoding use the legacy Flutter contract and are UTF-8 document text.
"""

import base64
import io
import logging

from app.schemas.chat import AttachmentIn

logger = logging.getLogger(__name__)

# Plain-text-ish file extensions accepted even when the client didn't send a
# "text/*" mime type (e.g. browsers often send "application/octet-stream" or
# omit the type for source files).
_PLAIN_TEXT_EXTENSIONS = (".txt", ".md", ".json", ".py", ".js", ".ts", ".dart", ".csv")

_SPREADSHEET_MIME_TYPES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/vnd.oasis.opendocument.spreadsheet",
)


def decode_attachment_bytes(attachment: AttachmentIn) -> bytes:
    """Decode an attachment payload without imposing an ASCII-only path.

    Explicit encodings are authoritative. Older Flutter clients omitted the
    field and sent document text as UTF-8, so an untagged ASCII string must not
    be guessed as base64 (values such as ``test`` are valid in both formats).
    Passing legacy Unicode text directly to ``b64decode`` produced the reported
    ``string argument should contain only ASCII characters`` failure.
    """
    if attachment.encoding == "utf-8":
        return attachment.data.encode("utf-8")
    if attachment.encoding == "base64":
        return base64.b64decode(attachment.data, validate=True)

    return attachment.data.encode("utf-8")


async def _extract_pdf_text(pdf_bytes: bytes) -> str:
    """Extract Unicode text from PDF bytes."""
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(pdf_bytes))
        text_pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_pages.append(text)
        return "\n".join(text_pages) if text_pages else "[PDF: no extractable text]"
    except Exception as e:
        logger.warning("PDF extraction failed: %s", e)
        return f"[PDF extraction error: {e}]"


async def _extract_csv_text(data: bytes, filename: str) -> str:
    """Extract and summarize CSV content."""
    text_content = data.decode("utf-8", errors="replace")
    lines = text_content.splitlines()
    row_count = len(lines)
    summary = f"[CSV: {filename}, {row_count} rows]\n"
    if row_count <= 50:
        summary += text_content
    else:
        summary += "--- Preview (first 50 rows) ---\n"
        summary += "\n".join(lines[:50])
        summary += f"\n... and {row_count - 50} more rows"
    return summary


async def _extract_spreadsheet_text(data: bytes, filename: str) -> str:
    """Extract and summarize Excel/openoffice spreadsheet content."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        summary_parts = []
        total_rows = 0
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            total_rows += len(rows)
            if rows:
                summary_parts.append(f"Sheet: {sheet_name}")
                if len(rows) <= 20:
                    for row in rows:
                        cells = [str(c) if c is not None else "" for c in row]
                        summary_parts.append(" | ".join(cells))
                else:
                    summary_parts.append(f"  {len(rows)} rows")
                    for row in rows[:10]:
                        cells = [str(c) if c is not None else "" for c in row]
                        summary_parts.append(" | ".join(cells))
                    summary_parts.append(f"  ... and {len(rows) - 10} more rows")
            summary_parts.append("")
        return f"[Spreadsheet: {filename}, {total_rows} total rows]\n" + "\n".join(summary_parts)
    except Exception as e:
        logger.warning("Spreadsheet extraction failed: %s", e)
        return f"[Spreadsheet extraction error: {e}]"


async def extract_attachment_text(attachment: AttachmentIn) -> str:
    """Return extracted text for a ``document``-type chat attachment.

    Dispatches on ``attachment.mime_type``:
      * ``application/pdf`` — extracted page text via pypdf
      * ``text/csv`` — decoded with a row-count header and 50-row preview
      * xlsx/xls/ods — sheet-by-sheet preview via openpyxl
      * any other ``text/*`` mime type, or a filename ending in one of
        ``.txt .md .json .py .js .ts .dart .csv`` — base64-decoded as UTF-8

    Any other mime type, or any extraction failure (including corrupt/
    invalid base64), falls back to returning ``attachment.data`` unchanged
    so a single bad attachment never raises out of message sending.
    """
    mime = attachment.mime_type.lower()
    filename = attachment.name
    data = attachment.data

    try:
        file_bytes = decode_attachment_bytes(attachment)
        if mime == "application/pdf":
            return await _extract_pdf_text(file_bytes)
        if mime == "text/csv":
            return await _extract_csv_text(file_bytes, filename)
        if mime in _SPREADSHEET_MIME_TYPES:
            return await _extract_spreadsheet_text(file_bytes, filename)
        if mime.startswith("text/") or filename.lower().endswith(_PLAIN_TEXT_EXTENSIONS):
            return file_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        logger.warning("Attachment text extraction failed for %s (%s): %s", filename, mime, e)
        return data

    # Unknown mime type: use the raw text as-is.
    return data
